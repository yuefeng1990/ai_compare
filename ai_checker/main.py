import sys
import os

# 设置标准输出编码，跟随当前终端编码，避免中文打印乱码
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding=sys.stdout.encoding or 'utf-8', errors='replace')
        sys.stderr.reconfigure(encoding=sys.stderr.encoding or 'utf-8', errors='replace')
    except AttributeError:
        import io
        stdout_encoding = sys.stdout.encoding or 'utf-8'
        stderr_encoding = sys.stderr.encoding or 'utf-8'
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding=stdout_encoding, errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding=stderr_encoding, errors='replace')

import pandas as pd
from tools.excel_reader import ExcelReader
from tools.web_checker import WebChecker
from tools.compare import CompareTool
from auto_login import automate_browser, automate_browser_with_search

def read_log_text(file_path):
    """读取日志文本，兼容 UTF-8 和常见中文 Windows 编码。"""
    candidates = []
    for index, encoding in enumerate(('utf-8-sig', 'utf-8', 'gb18030', 'gbk')):
        try:
            with open(file_path, 'r', encoding=encoding, errors='strict') as f:
                content = f.read()
        except UnicodeDecodeError:
            continue

        candidates.append((_mojibake_score(content), index, content, encoding))

    if candidates:
        _, _, content, encoding = min(candidates, key=lambda item: (item[0], item[1]))
        return content, encoding

    with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
        return f.read(), 'utf-8(errors=replace)'


def _mojibake_score(text):
    """给解码结果打分，分数越高越像乱码。"""
    score = 0
    score += text.count('\ufffd') * 100
    score += text.count('ï¿½') * 100
    score += text.count('锟斤拷') * 80
    score += text.count('锟') * 20
    score += sum(1 for ch in text if '\u0370' <= ch <= '\u03ff') * 5
    score += sum(1 for ch in text if '\u0400' <= ch <= '\u04ff') * 5
    return score


def has_mojibake_markers(text):
    """判断日志内容是否已经包含常见乱码标记。"""
    return any(marker in text for marker in ('\ufffd', 'ï¿½', '锟斤拷', '锟'))


def is_os_version_greater_than(os_full_name, minimum_version='5.0'):
    """判断 OsFullName 中的主/次版本是否大于指定版本。"""
    import re

    if not os_full_name:
        return False, None

    match = re.search(r'(\d+)\.(\d+)', str(os_full_name))
    if not match:
        return False, None

    version_tuple = (int(match.group(1)), int(match.group(2)))
    min_major, min_minor = (int(part) for part in minimum_version.split('.', 1))
    return version_tuple > (min_major, min_minor), f"{version_tuple[0]}.{version_tuple[1]}"


def extract_keywords_from_log(log_content):
    """
    从log.txt内容中提取关键字（key = value 或 key: value 格式中的key）。
    如果存在 Product Params 标记，只提取 Start 和 End 之间的关键字。
    支持中文关键字，并会去掉开头的 get 前缀。
    """
    import re
    keywords = []
    in_params = False
    has_params_marker = 'To Obtain Product Params Start' in log_content
    allowed_key_pattern = re.compile(r'^[a-zA-Z0-9_\s\u4e00-\u9fff（）()\-]+$')
    seen = set()

    for line in log_content.split('\n'):
        line_stripped = line.strip()

        # 检测参数段开始
        if 'To Obtain Product Params Start' in line_stripped:
            in_params = True
            continue

        # 检测参数段结束
        if 'To Obtain Product Params End' in line_stripped:
            break

        if has_params_marker and not in_params:
            continue

        # 跳过空行
        if not line_stripped:
            continue

        # 提取 key = value / key: value 格式中的key
        separator = '=' if '=' in line_stripped else ':' if ':' in line_stripped else None
        if separator:
            key_part = line_stripped.split(separator, 1)[0].strip()
            key_part = re.sub(r'^\s*get[\s_-]*', '', key_part, flags=re.IGNORECASE).strip()
            normalized_key = CompareTool._normalize_keyword(key_part)

            # 只保留干净的参数名：中英文、数字、空白和常见连接符/括号
            if key_part and allowed_key_pattern.match(key_part) and normalized_key not in seen:
                keywords.append(key_part)
                seen.add(normalized_key)

    return keywords


def main(measurement_id=None):
    """
    主函数
    :param measurement_id: 测评编号，如果提供则自动搜索该编号的详情信息
    """
    # 文件路径设置
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_excel_path = os.path.join(current_dir, 'data', 'checklist.xlsx')
    keywords_txt_path = os.path.join(current_dir, 'data', 'keywords.txt')
    log_txt_path = os.path.join(current_dir, 'data', 'test.txt')
    output_excel_path = os.path.join(current_dir, 'output', 'result.xlsx')
    
    # 网站URL定义
    WEB_URL = "https://compatibility.openharmony.cn/mng/index"
    
    web_checker = None
    try:
        # 1. 读取log.txt内容
        print("1. 读取log.txt内容...")
        log_content = ""
        if os.path.exists(log_txt_path):
            log_content, log_encoding = read_log_text(log_txt_path)
            print(f"log.txt文件大小: {len(log_content)} 字符，编码: {log_encoding}")
            if has_mojibake_markers(log_content):
                print("⚠ 日志内容仍包含乱码标记（如 � / 锟斤拷），请确认源日志未被错误编码保存。")
        else:
            print(f"警告: {log_txt_path} 不存在")
            return

        # 2. 从log.txt提取关键字
        print("\n2. 从log.txt提取关键字...")
        all_keywords = extract_keywords_from_log(log_content)
        print(f"从log.txt提取到 {len(all_keywords)} 个关键字")
        for kw in all_keywords:
            print(f"    - {kw}")

        # 保存关键字到 keyword.txt
        print(f"\n   保存关键字到 {keywords_txt_path} ...")
        with open(keywords_txt_path, 'w', encoding='utf-8') as f:
            for kw in all_keywords:
                f.write(kw + '\n')
        print(f"   ✓ 已保存 {len(all_keywords)} 个关键字到 keyword.txt")

        # 使用9个目标关键字作为检查清单（log提取值通过归一化匹配）
        TARGET_KEYWORDS = [
            'Manufacture', 'OsFullName', 'MarketName', 'ProductModel',
            'Brand', 'DisplayVersion', 'VersionId',
            'SecurityPatchTag', 'BuildRootHash'
        ]
        print(f"\n   目标关键字（共 {len(TARGET_KEYWORDS)} 个）:")
        for kw in TARGET_KEYWORDS:
            print(f"      - {kw}")
        all_keywords = TARGET_KEYWORDS

        # 3. 读取Excel检查清单（参考用）
        print("\n3. 读取Excel检查清单...")
        excel_reader = ExcelReader(input_excel_path)
        checklist_data = excel_reader.read_excel()
        
        if checklist_data is None or checklist_data.empty:
            print("未找到检查清单数据！")
            return
        
        print(f"检查清单包含 {len(checklist_data)} 行数据")
        
        # 4. 初始化对比工具（传入目标测评编号用于精确定位表格行）
        print("\n4. 初始化对比工具...")
        compare_tool = CompareTool(target_measurement_id=measurement_id)
        
        # 5. 如果提供了测评编号，启动浏览器自动化
        web_content = ""
        if measurement_id:
            print(f"\n5. 启动浏览器自动化，搜索测评编号: {measurement_id}")
            print("=" * 80)
            
            # 每次都启动浏览器获取最新页面内容（不使用缓存）
            print("⚠ 将启动浏览器获取最新页面内容...")
            
            # 初始化WebChecker
            from tools.web_checker import WebChecker
            import time
            web_checker = WebChecker(headless=True)  # 可视化模式,False表示可视化，True表示无头模式
            
            try:
                # 启动浏览器
                print("      - 正在启动浏览器...")
                web_checker.launch_browser()
                print("      ✓ 浏览器已启动")
                
                # 使用带搜索功能的自动化脚本
                success = automate_browser_with_search(web_checker, measurement_id)
                
                if not success:
                    print("\n✗ 浏览器自动化流程失败")
                    web_content = ""
                else:
                    # 获取页面内容
                    print("\n      - 获取页面内容...")
                    
                    # 优先使用detail_frame（详情页的iframe），如果不存在则使用page
                    target_frame = None
                    if hasattr(web_checker, 'detail_frame') and web_checker.detail_frame:
                        target_frame = web_checker.detail_frame
                        print("      ✓ 使用详情页iframe获取内容")
                    elif web_checker and web_checker.page:
                        target_frame = web_checker.page
                        print("      ⚠ 使用主页面获取内容")
                    
                    if target_frame:
                        try:
                            # 滚动页面触发懒加载
                            target_frame.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                            time.sleep(2)
                            
                            # 获取页面HTML内容（用于成对div结构提取）
                            page_html = target_frame.content()
                            
                            # 同时获取纯文本作为备选
                            page_content = target_frame.inner_text('body')
                            
                            # 优先使用HTML内容，如果为空则使用纯文本
                            if page_html and len(page_html) > 100:
                                web_content = page_html
                                print(f"\n✓ 成功获取页面HTML内容: {len(web_content)} 字符")
                            elif page_content:
                                web_content = page_content
                                print(f"\n✓ 成功获取页面文本内容: {len(web_content)} 字符")
                            else:
                                print("\n✗ 未能获取到页面内容")
                                web_content = ""
                            
                            # 保存页面内容用于调试
                            if web_content:
                                debug_file = os.path.join(current_dir, 'page_content_debug.txt')
                                with open(debug_file, 'w', encoding='utf-8') as f:
                                    f.write(web_content)
                                print(f"✓ 页面内容已保存到: {debug_file}")
                        except Exception as e:
                            print(f"\n✗ 获取页面内容失败: {str(e)}")
                    else:
                        print("\n✗ 浏览器对象无效，无法获取页面内容")
                        
            except Exception as e:
                print(f"\n✗ 浏览器自动化异常: {str(e)}")
                print("⚠ 将跳过网页提取，继续处理其他数据源")
                web_content = ""
            finally:
                # 关闭浏览器
                try:
                    web_checker.close()
                    print("✓ 浏览器已关闭")
                except:
                    pass
        else:
            print("\n5. 跳过浏览器自动化（未提供测评编号）")
        
        # 6. 从log.txt提取关键字值
        print("\n6. 从log.txt提取关键字...")
        print("=" * 80)
        
        log_values = {}
        for keyword in all_keywords:
            value = compare_tool.extract_value_from_log(log_content, keyword)
            log_values[keyword] = value
            status = "✓" if value else "✗"
            print(f"  {status} {keyword:20s}: {value or '(未找到)'}")
        
        # 7. 从网页提取关键字值（如果有网页内容）
        web_values = {}
        if web_content:
            print("\n7. 从网页提取关键字...")
            print("=" * 80)
            
            for keyword in all_keywords:
                value = compare_tool.extract_value_from_web(web_content, keyword)
                web_values[keyword] = value
                status = "✓" if value else "✗"
                print(f"  {status} {keyword:20s}: {value or '(未找到)'}")
        else:
            print("\n7. 跳过网页提取（无网页内容）")
        
        # 8. 比较log.txt和网页的值
        print("\n8. 比较结果...")
        print("=" * 80)
        
        comparison_results = []
        matched_count = 0
        total_count = len(all_keywords)
        
        for keyword in all_keywords:
            log_value = log_values.get(keyword)
            web_value = web_values.get(keyword)
            
            is_match = compare_tool.compare_values(log_value, web_value, keyword=keyword)
            
            if is_match:
                matched_count += 1
            
            status = "✓" if is_match else "✗"
            match_text = "一致" if is_match else "不一致"
            
            print(f"\n  {status} {keyword:20s}")
            print(f"      log.txt: {log_value or '(未找到)'}")
            print(f"      网页:    {web_value or '(未找到)'}")
            print(f"      结果:    {match_text}")
            
            comparison_results.append({
                '关键字': keyword,
                'log값': log_value or '',
                'web값': web_value or '',
                '是否一致': '✓' if is_match else '❌'
            })
        
        # 9. 生成结果Excel
        print("\n" + "=" * 80)
        print("9. 生成结果Excel...")
        
        # 创建输出目录
        os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
        
        # 创建DataFrame并保存
        df_result = pd.DataFrame(comparison_results)
        df_result.to_excel(output_excel_path, index=False)
        
        print(f"✓ 结果已保存到: {output_excel_path}")
        
        # 10. 将一致的关键字数据写入data/checklist.xlsx
        print("\n10. 将一致的关键字数据写入checklist.xlsx...")
        print("=" * 80)
        
        from openpyxl import load_workbook
        
        # 关键字（归一化）-> checklist C列匹配文本 映射
        # 统一使用 CompareTool._normalize_keyword()，忽略大小写、空格和开头 get
        _norm = CompareTool._normalize_keyword
        keyword_checklist_map = {
            _norm('MarketName'): '设备名称（传播名）',
            _norm('ProductModel'): '设备型号',
            _norm('Manufacture'): '企业简称（英文）',
            _norm('DisplayVersion'): '软件版本号',
            _norm('SecurityPatchTag'): '安全补丁标签',
            _norm('VersionId'): '版本id',
            _norm('BuildRootHash'): '版本Hash',
            _norm('OsFullName'): '操作系统版本号',
            _norm('Brand'): '品牌英文名称',
        }

        wb = load_workbook(input_excel_path)
        ws = wb.active

        # 先清空F列（第6列，索引5）所有值
        print("   清空F列现有值...")
        cleared_count = 0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            if row[5].value is not None:
                row[5].value = None
                cleared_count += 1
        print(f"   ✓ 已清空 {cleared_count} 个F列单元格")

        written_count = 0

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            seq = row[0].value or ''
            if row[0].row == 5:
                os_full_name = log_values.get('OsFullName')
                is_version_ok, os_version = is_os_version_greater_than(os_full_name, '5.0')
                row[5].value = "✔" if is_version_ok else "❌"
                written_count += 1
                version_text = os_version or '未提取到版本'
                print(f"  {'✓' if is_version_ok else '❌'} [序号{seq}] OsFullName版本 {version_text} > 5.0: 已写入结果")
                continue

            c_value = row[2].value  # C列 - 测试检查项
            if c_value is None:
                continue

            c_text = str(c_value)

            for keyword, search_text in keyword_checklist_map.items():
                if search_text in c_text:
                    # 检查该关键字（归一化后）是否匹配一致
                    result_item = next(
                        (item for item in comparison_results if _norm(item['关键字']) == keyword),
                        None
                    )
                    if result_item:
                        if result_item['是否一致'] == '✓':
                            row[5].value = f"✔"
                            written_count += 1
                            print(f"  ✓ [序号{seq}] {result_item['关键字']}: 已写入一致结果")
                        elif result_item['是否一致'] == '❌':
                            row[5].value = f"❌"
                            written_count += 1
                            print(f"  ❌ [序号{seq}] {result_item['关键字']}: 已写入不一致结果")
                    break
        
        wb.save(input_excel_path)
        print(f"  ✓ 共 {written_count} 项结果已写入checklist.xlsx")
        
        # 统计一致性
        print(f"\n统计信息:")
        print(f"  总比较次数: {total_count}")
        print(f"  一致次数: {matched_count}")
        print(f"  不一致次数: {total_count - matched_count}")
        if total_count > 0:
            print(f"  一致率: {matched_count/total_count*100:.2f}%")
        else:
            print(f"  一致率: N/A")
        
        print("\n" + "=" * 80)
        print("✅ 处理完成！")
        print("=" * 80)
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭浏览器（如果还在运行）
        if web_checker:
            try:
                web_checker.close()
                print("\n浏览器已关闭")
            except:
                pass

if __name__ == "__main__":
    # 从命令行参数获取测评编号
    measurement_id = sys.argv[1] if len(sys.argv) > 1 else None
    
    if measurement_id:
        print(f"📋 测评编号: {measurement_id}")
    else:
        print("⚠️  未提供测评编号，将跳过浏览器自动化")
    
    main(measurement_id)
